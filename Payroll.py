import openpyxl
import uuid

path = 'Payroll.xlsx'


class Employee:
    def __init__(self, name = '', position = '', salary = 0.00, exp = 0, ID = None):
        self.name = name
        self.position = position
        self.salary = salary
        self.exp = exp
        self.ID = ID
    
    
    def getname(self): return self.name 
    def getposition(self): return self.position
    def getsalary(self): return self.salary
    def getexp(self): return self.exp
    def getID(self): return self.ID


def createID():
    return uuid.uuid4()


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws['A1'] = 'Employee'
    ws['B1'] = 'Position'
    ws['C1'] = 'Salary(per Year)'
    ws['D1'] = 'Experience'
    ws['E1'] = 'ID'
    ws['G1'] = 'Salary Cap'
    ws['H1'] = 'Payroll'
    
    print('workbook created for Payroll')
    wb.save(path)


def create_employee(employee_name = '', position = '', salary = 0.00):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    employee = Employee(employee_name, position, salary, 0, createID())
    
    r = 2
    while True:
        print('checking...')
        if ws.cell(row=r, column=1).value == None:
            ws.cell(row=r, column=1).value = employee.getname()
            ws.cell(row=r, column=2).value = employee.getposition()
            ws.cell(row=r, column=3).value = employee.getsalary()
            ws.cell(row=r, column=4).value = employee.getexp()
            ws.cell(row=r, column=5).value = str(employee.getID())
            wb.save(path)
            break
        else:
            r += 1
        
                   
def find_employee(id):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    c = 5
    r = 2
    while True:
        if ws.cell(row=r, column = c).value == id:
            print('getting info...')
            
            #get employee info
            employee = Employee(ws.cell(row=r, column=1).value,
                                ws.cell(row=r, column=2).value,
                                ws.cell(row=r, column=3).value,
                                ws.cell(row=r, column=4).value,
                                uuid.UUID(id))
            
            return employee
            
        elif ws.cell(row=r, column = c).value == None:
            print(f'Employee with ID {uuid} does not exist')
            break
        else:
            print('Something went wrong!')
            break


